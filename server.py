from flask import Flask, jsonify, request
from flask_cors import CORS
import requests
import json
import random
from datetime import datetime, timedelta
from openpyxl import load_workbook  # Import openpyxl

app = Flask(__name__)
CORS(app)

WEATHERAPI_KEY = "91505f7efd4b494eb66191358240509"  # Replace with your API key!
CACHE_FILE = 'weather_cache.json'
CACHE_DURATION = timedelta(hours=1)  # Cache duration set to 1 hour

def load_city_data():
    print("Loading city data from palestine_cities.xlsx...")
    try:
        workbook = load_workbook('palestine_cities.xlsx')
        sheet = workbook.active

        cities = {}
        for row in sheet.iter_rows(min_row=2):  # Start from row 2, assuming row 1 is header
            city_name = row[0].value  # Column A
            city_name_ar = row[1].value # Column B
            land_area = row[2].value  # Column C
            population = row[3].value # Column D
            latitude = row[4].value   # Column E
            longitude = row[5].value  # Column F
            population_density = row[6].value # Column G
            coordinates_str = row[7].value  # Column H (Coordinates string)

            # Extract coordinates from the string 
            if coordinates_str:
                coordinates = [float(coord) for coord in coordinates_str.split(',')] 
            else:
                coordinates = [None, None] # Handle cases where coordinates are missing

            city_name_lower = city_name.lower()
            cities[city_name_lower] = {
                "city": city_name,
                "location_ar": city_name_ar,
                "land_area": land_area,
                "population": population,
                "coordinates": coordinates,  
                "population_density": population_density
            }

        print("City data loaded successfully.")
        return cities

    except FileNotFoundError:
        print("Error: palestine_cities.xlsx not found.")
        return {}
    except Exception as e:
        print(f"Error loading data from Excel file: {e}")
        return {}

# Load weather data from cache
def load_cache():
    print("Loading weather cache from:", CACHE_FILE)
    try:
        with open(CACHE_FILE, 'r') as f:
            cache = json.load(f)
            print("Cache loaded successfully.")
            return cache
    except FileNotFoundError:
        print("Cache file not found. Initializing new cache.")
        return {'timestamp': None, 'data': {}}
    except json.JSONDecodeError:
        print("Error decoding cache JSON. Initializing new cache.")
        return {'timestamp': None, 'data': {}}

# Save weather data to cache
def save_cache(cache):
    print("Saving weather cache to:", CACHE_FILE)
    try:
        with open(CACHE_FILE, 'w') as f:
            json.dump(cache, f)
        print("Cache saved successfully.")
    except Exception as e:
        print(f"Error saving cache: {e}")

# Fetch weather data using WeatherAPI (modified)
def fetch_weather_data(city_name, coordinates=None, api_key=WEATHERAPI_KEY):
    print(f"Fetching weather data for {city_name} from WeatherAPI...")
    url = "http://api.weatherapi.com/v1/forecast.json"
    params = {
        "key": api_key,
        "days": 1,  
        "aqi": "no"
    }

    if coordinates:
        params["q"] = f"{coordinates[1]},{coordinates[0]}"  # Latitude, Longitude
    else:
        params["q"] = city_name 

    try:
        response = requests.get(url, params=params)
        response.raise_for_status()
        data = response.json()

        weather_info = {
            "current": {
                "temperature": data['current']['temp_c'],
                "humidity": data['current']['humidity'],
                "wind_speed": data['current']['wind_kph'],
                "precipitation": data['current']['precip_mm'],
                "description": data['current']['condition']['text']
            },
            "hourly": []
        }

        for hour_data in data['forecast']['forecastday'][0]['hour']:
            weather_info["hourly"].append({
                "temperature": hour_data['temp_c'],
                "humidity": hour_data['humidity'],
                "wind_speed": hour_data['wind_kph'],
                "precipitation": hour_data['precip_mm'],
                "description": hour_data['condition']['text']
            })

        print(f"Weather data fetched successfully for {city_name}")
        return weather_info

    except requests.exceptions.RequestException as e:
        print(f"Error fetching weather data: {e}")
        return None


def get_or_update_weather_cache(city_name, coordinates=None):
    cache = load_cache()
    now = datetime.now()

    # Create a unique cache key based on city name or coordinates 
    cache_key = city_name 
    if coordinates:
        cache_key += f"_{coordinates[1]}_{coordinates[0]}"

    # Check if data in cache and not expired
    if cache_key in cache['data'] and now - datetime.fromisoformat(cache['data'][cache_key]['timestamp']) < CACHE_DURATION:
        print(f"Using cached weather data for {city_name}")
        return cache['data'][cache_key]

    # Fetch new weather data (using coordinates if provided)
    weather_data = fetch_weather_data(city_name, coordinates=coordinates) 
    if weather_data:
        cache['data'][cache_key] = {'weather': weather_data, 'timestamp': now.isoformat()}
        save_cache(cache)
        return cache['data'][cache_key]
    return None


@app.route('/api/cities', methods=['GET'])
def get_cities():
    cities = load_city_data()
    sorted_cities = sorted(list(cities.keys())) # Sort city names alphabetically
    return jsonify(sorted_cities) 

@app.route('/api/city/<city>', methods=['GET'])
def get_city_data(city):
    city_name = city.lower()
    cities = load_city_data()
    if city_name in cities:
        city_info = cities[city_name]
        weather_data = get_or_update_weather_cache(city_name)

        if not weather_data:  # If fetching by name fails
            coordinates = city_info.get('coordinates') 
            if coordinates:
                print(f"Trying to fetch weather using coordinates for {city_name}")
                weather_data = get_or_update_weather_cache(city_name, coordinates=coordinates) 

        if weather_data:
            city_info.update({'weather': weather_data['weather']})
            city_info.update({'weather_timestamp': weather_data['timestamp']})
        return jsonify(city_info)
    else:
        return jsonify({"error": "City not found"}), 404

@app.route('/api/predict', methods=['POST'])
def predict():
    data = request.get_json()
    city_name = data.get('city', '').lower()
    print(f"Prediction requested for city: {city_name}")
    if city_name:
        prediction_result = random.random()  
    else:
        prediction_result = "Error: City name not provided in the request."
    return jsonify({"prediction": prediction_result})


if __name__ == '__main__':
    app.run(debug=True)